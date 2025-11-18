"""
Centralised styling for the R&R dashboard.

This module defines styling functions and utilities including CSS injection
and Excel export capabilities. By consolidating fonts, colours, component
styling and export functions into one place, we ensure a uniform look and
feel across all pages.

Usage:

```python
import styles
styles.apply_styles()

# For Excel export
excel_file = styles.create_excel_download(df, "report.xlsx")
st.download_button(
    label="📊 Download Excel",
    data=excel_file,
    file_name="report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
```

This should be called at the top of each page or main function before
rendering any UI elements.
"""
import streamlit as st
import pandas as pd
from io import BytesIO
from typing import Dict, Optional


def apply_styles(theme: str = "blue"):
    """Inject global CSS for the dashboard with a selectable colour theme.

    Parameters
    ----------
    theme : str, optional
        Name of the colour theme to use.  Supported values are
        "blue", "yellow", and "white" (case‑insensitive).  The
        default is "blue".

    This function generates CSS rules on the fly based on the chosen
    theme.  It also updates Plotly's default colour palette for
    discrete charts to ensure consistent visuals across the dashboard.
    The CSS includes definitions for the page background, metric cards,
    chart containers and navigation bar.  It should be called at the
    start of every page after any theme selection is made.
    """
    import plotly.express as px
    theme_key = theme.lower() if theme else "blue"
    # Define colour palettes for supported themes
    THEME_DEFINITIONS = {
        "blue": {
            "gradient": ["#0D3C66", "#0979B7", "#4FC5FA"],
            "accent_dark": "#0D3C66",
            "accent_base": "#0979B7",
            "accent_light": "#4FC5FA",
            "text_primary": "#222222",
            "nav_fg": "#0D3C66",
            "chart_palette": ["#4FC5FA", "#0979B7", "#0D3C66", "#00BFB2", "#F5B041", "#E74C3C", "#58D68D"]
        },
        "yellow": {
            "gradient": ["#F4C542", "#F8A13B", "#FCE38A"],
            "accent_dark": "#BE6515",
            "accent_base": "#D97D20",
            "accent_light": "#F8A13B",
            "text_primary": "#222222",
            "nav_fg": "#BE6515",
            "chart_palette": ["#F4C542", "#F8A13B", "#EF6B3F", "#8BC34A", "#00B0FF", "#E91E63", "#9C27B0"]
        },
        "white": {
            "gradient": ["#F5F5F5", "#F0F0F0", "#EAEAEA"],
            "accent_dark": "#333333",
            "accent_base": "#555555",
            "accent_light": "#888888",
            "text_primary": "#222222",
            "nav_fg": "#333333",
            "chart_palette": ["#0D3C66", "#0979B7", "#4FC5FA", "#F4C542", "#EF6B3F", "#8BC34A", "#E91E63"]
        }
    }
    colours = THEME_DEFINITIONS.get(theme_key, THEME_DEFINITIONS["blue"])
    # Update Plotly default discrete colour sequence for bright, high‑contrast charts
    px.defaults.color_discrete_sequence = colours["chart_palette"]
    # Build the CSS string with theme variables
    gradient_css = f"linear-gradient(120deg, {colours['gradient'][0]} 0%, {colours['gradient'][1]} 50%, {colours['gradient'][2]} 100%)"
    nav_bg = "rgba(255, 255, 255, 0.5)" if theme_key != "white" else "rgba(255, 255, 255, 0.8)"
    button_bg = "rgba(255, 255, 255, 0.3)"
    button_hover_bg = "rgba(255, 255, 255, 0.4)"
    css = f"""
    <style>
    /* Base page settings */
    body {{
        background: {gradient_css};
        color: {colours['text_primary']};
        font-family: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        min-height: 100vh;
        margin: 0;
    }}

    /* Ensure the main container has minimal top padding */
    .main > div {{
        padding-top: 0.5rem;
    }}

    /* Section titles */
    .section-title {{
        font-weight: 600;
        font-size: 1.05rem;
        color: {colours['accent_dark']};
        margin-bottom: 0.4rem;
    }}

    /* KPI card styling */
    .metric-card {{
        background: rgba(255, 255, 255, 0.85);
        border-radius: 14px;
        padding: 14px 16px;
        border: 1px solid rgba(255, 255, 255, 0.35);
        box-shadow: 0 10px 25px rgba(15, 23, 42, 0.15);
        text-align: center;
        backdrop-filter: blur(6px);
        -webkit-backdrop-filter: blur(6px);
    }}
    .metric-card h4 {{
        font-size: 0.9rem;
        font-weight: 600;
        color: {colours['accent_base']};
        margin-bottom: 0.25rem;
    }}
    .metric-card h2 {{
        font-size: 1.6rem;
        font-weight: 700;
        color: {colours['accent_dark']};
        margin: 0;
    }}

    /* KPI label and helper */
    .kpi-label {{
        display: inline-flex;
        align-items: center;
        justify-content: center;
        gap: 4px;
        white-space: nowrap;
    }}
    .kpi-help {{
        font-size: 0.75rem;
        color: #6b7280;
        border-radius: 999px;
        border: 1px solid #d4d4d8;
        width: 16px;
        height: 16px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        cursor: help;
        background-color: #f9fafb;
        position: relative;
    }}
    .kpi-help::after {{
        content: attr(data-tip);
        position: absolute;
        left: 50%;
        bottom: 125%;
        transform: translateX(-50%);
        background: {colours['accent_dark']};
        color: #f9fafb;
        padding: 6px 10px;
        border-radius: 6px;
        font-size: 0.75rem;
        line-height: 1.2;
        white-space: normal;
        min-width: 180px;
        max-width: 260px;
        text-align: left;
        opacity: 0;
        pointer-events: none;
        transition: opacity 0.15s ease-out;
        z-index: 9999;
    }}
    .kpi-help::before {{
        content: "";
        position: absolute;
        left: 50%;
        bottom: 115%;
        transform: translateX(-50%);
        border-width: 5px;
        border-style: solid;
        border-color: {colours['accent_dark']} transparent transparent transparent;
        opacity: 0;
        transition: opacity 0.15s ease-out;
    }}
    .kpi-help:hover::after,
    .kpi-help:hover::before {{
        opacity: 1;
    }}

    /* Plotly chart container styling */
    .stPlotlyChart {{
        background: rgba(255, 255, 255, 0.85);
        border-radius: 16px;
        padding: 12px;
        margin: 0.5rem 0 1.5rem 0;
        border: 1px solid rgba(255, 255, 255, 0.35);
        box-shadow: 0 10px 30px rgba(15, 23, 42, 0.15);
        backdrop-filter: blur(6px);
        -webkit-backdrop-filter: blur(6px);
    }}

    /* Navigation bar styling */
    .navbar-wrapper {{
        padding: 0.6rem 0.2rem 0.8rem 0.2rem;
        background: {nav_bg};
        border-bottom: 1px solid rgba(255, 255, 255, 0.3);
        position: relative;
        backdrop-filter: blur(6px);
        -webkit-backdrop-filter: blur(6px);
        z-index: 999;
    }}
    .stButton > button {{
        transition: all 0.2s ease-in-out;
        border-radius: 8px;
        font-weight: 500;
        border: 1px solid {colours['accent_light']};
        background-color: {button_bg};
        color: {colours['nav_fg']};
    }}
    .stButton > button:hover {{
        transform: translateY(-2px);
        background-color: {button_hover_bg};
    }}
    
    /* Download button styling */
    .stDownloadButton > button {{
        background-color: {colours['accent_base']};
        color: white;
        font-weight: 600;
        border-radius: 8px;
        padding: 0.5rem 1.5rem;
        border: none;
        transition: all 0.2s ease-in-out;
    }}
    .stDownloadButton > button:hover {{
        background-color: {colours['accent_dark']};
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
    }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)
    

def show_spinner(duration: float = 1.0):
    """
    Display a custom loading spinner with the Acies colour palette.

    A semi‑transparent circular ring animates around a trophy icon
    (implemented via an emoji) to indicate loading.  The ring
    continuously rotates for the specified ``duration`` (in
    seconds), after which it disappears.  Use this at the top of a
    page or before heavy computations to signal that content is
    loading.

    Parameters
    ----------
    duration : float
        Length of time to show the spinner, in seconds.  Defaults
        to 1.0.  Adjust as needed based on expected loading time.
    """
    import time
    # Create a placeholder so we can remove the spinner afterwards
    placeholder = st.empty()
    spinner_html = """
    <div class="rr-loading">
      <div class="rr-spinner"></div>
      <span class="rr-icon">🏆</span>
    </div>
    <style>
    .rr-loading {
      position: fixed;
      top: 50%;
      left: 50%;
      transform: translate(-50%, -50%);
      display: flex;
      justify-content: center;
      align-items: center;
      width: 160px;
      height: 160px;
      z-index: 10000;
      pointer-events: none;
    }
    .rr-spinner {
      position: absolute;
      width: 120px;
      height: 120px;
      border: 8px solid rgba(255,255,255,0.3);
      border-top-color: #4FC5FA;
      border-right-color: #0979B7;
      border-bottom-color: #0D3C66;
      border-left-color: #4FC5FA;
      border-radius: 50%;
      animation: rr-spin 1.5s linear infinite;
    }
    .rr-icon {
      font-size: 48px;
      color: #ffffff;
      z-index: 1;
    }
    @keyframes rr-spin {
      to { transform: rotate(360deg); }
    }
    </style>
    """
    # Render spinner
    placeholder.markdown(spinner_html, unsafe_allow_html=True)
    # Keep spinner on screen for the specified duration
    time.sleep(max(duration, 0.1))
    # Clear the spinner
    placeholder.empty()


# ============================================================================
# EXCEL EXPORT FUNCTIONS
# ============================================================================

def create_excel_download(
    dataframe: pd.DataFrame,
    filename: str = "data_export.xlsx",
    sheet_name: str = "Data"
) -> BytesIO:
    """
    Create an Excel file from a DataFrame for download.
    
    This function converts a pandas DataFrame into an Excel (.xlsx) file
    stored in memory, ready to be used with Streamlit's download_button.
    The Excel file is created using openpyxl engine for maximum compatibility.
    
    Parameters
    ----------
    dataframe : pd.DataFrame
        The data to export to Excel format
    filename : str, optional
        Name of the downloaded file (default: "data_export.xlsx")
    sheet_name : str, optional
        Name of the Excel sheet (default: "Data")
        
    Returns
    -------
    BytesIO
        Excel file buffer ready for st.download_button
        
    Example
    -------
    >>> excel_file = create_excel_download(df, "sales_report.xlsx", "Q1 Sales")
    >>> st.download_button(
    ...     label="📥 Download Excel",
    ...     data=excel_file,
    ...     file_name="sales_report.xlsx",
    ...     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ... )
    """
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error creating Excel file: {str(e)}")
        return BytesIO()


def create_multi_sheet_excel(
    data_dict: Dict[str, pd.DataFrame],
    filename: str = "report.xlsx"
) -> BytesIO:
    """
    Export multiple DataFrames to different sheets in one Excel file.
    
    This function is useful for creating comprehensive reports where
    different data categories are organized into separate sheets within
    a single Excel workbook.
    
    Parameters
    ----------
    data_dict : dict
        Dictionary with {sheet_name: dataframe} pairs. Each key becomes
        a sheet name and its value is the DataFrame for that sheet.
    filename : str, optional
        Name of the downloaded file (default: "report.xlsx")
        
    Returns
    -------
    BytesIO
        Excel file buffer with multiple sheets ready for download
        
    Example
    -------
    >>> data = {
    ...     "Sales": sales_df,
    ...     "Inventory": inventory_df,
    ...     "Customers": customers_df
    ... }
    >>> excel_file = create_multi_sheet_excel(data, "monthly_report.xlsx")
    >>> st.download_button(
    ...     label="📊 Download Full Report",
    ...     data=excel_file,
    ...     file_name="monthly_report.xlsx",
    ...     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ... )
    """
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in data_dict.items():
                # Clean sheet name (Excel has 31 char limit and some forbidden chars)
                clean_name = sheet_name[:31].replace('/', '-').replace('\\', '-')
                df.to_excel(writer, sheet_name=clean_name, index=False)
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error creating multi-sheet Excel file: {str(e)}")
        return BytesIO()


def create_styled_excel(
    dataframe: pd.DataFrame,
    filename: str = "styled_export.xlsx",
    sheet_name: str = "Data",
    header_color: str = "0979B7",
    theme: str = "blue"
) -> BytesIO:
    """
    Create a professionally styled Excel file with formatted headers.
    
    This function creates an Excel file with styled headers matching your
    dashboard theme, including colored backgrounds, bold white text, and
    centered alignment. This makes the exported data look professional
    and consistent with your dashboard design.
    
    Parameters
    ----------
    dataframe : pd.DataFrame
        The data to export
    filename : str, optional
        Name of the downloaded file (default: "styled_export.xlsx")
    sheet_name : str, optional
        Name of the Excel sheet (default: "Data")
    header_color : str, optional
        Hex color for header background without # (default: "0979B7")
    theme : str, optional
        Theme name to match dashboard colors (default: "blue")
        
    Returns
    -------
    BytesIO
        Styled Excel file buffer ready for download
        
    Example
    -------
    >>> excel_file = create_styled_excel(
    ...     df, 
    ...     "styled_report.xlsx",
    ...     "Q4 Results",
    ...     theme="blue"
    ... )
    >>> st.download_button(
    ...     label="📥 Download Styled Report",
    ...     data=excel_file,
    ...     file_name="styled_report.xlsx",
    ...     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ... )
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Theme-based header colors
    THEME_COLORS = {
        "blue": "0979B7",
        "yellow": "D97D20",
        "white": "555555"
    }
    
    theme_key = theme.lower()
    header_bg_color = THEME_COLORS.get(theme_key, "0979B7")
    
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            # Style headers
            header_fill = PatternFill(
                start_color=header_bg_color,
                end_color=header_bg_color,
                fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column
            ):
                for cell in row:
                    cell.border = thin_border
                    # Center align data cells
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error creating styled Excel file: {str(e)}")
        return BytesIO()


def create_download_section(
    dataframe: pd.DataFrame,
    filename: str = "dashboard_export.xlsx",
    button_label: str = "📥 Download as Excel",
    sheet_name: str = "Data",
    use_styling: bool = True,
    theme: str = "blue"
):
    """
    Create a ready-to-use download section with a styled button.
    
    This is a convenience function that combines Excel creation and
    download button in one call. It handles all the boilerplate code
    for you and creates a professionally styled download experience.
    
    Parameters
    ----------
    dataframe : pd.DataFrame
        The data to export
    filename : str, optional
        Name of the downloaded file (default: "dashboard_export.xlsx")
    button_label : str, optional
        Text to display on download button (default: "📥 Download as Excel")
    sheet_name : str, optional
        Name of the Excel sheet (default: "Data")
    use_styling : bool, optional
        Whether to apply professional styling to Excel (default: True)
    theme : str, optional
        Theme for styling (default: "blue")
        
    Example
    -------
    >>> # Simple usage
    >>> create_download_section(df, "sales_data.xlsx")
    >>> 
    >>> # Custom usage
    >>> create_download_section(
    ...     df,
    ...     filename="Q4_report.xlsx",
    ...     button_label="📊 Export Q4 Data",
    ...     sheet_name="Q4 Sales",
    ...     use_styling=True,
    ...     theme="yellow"
    ... )
    """
    if use_styling:
        excel_file = create_styled_excel(
            dataframe,
            filename=filename,
            sheet_name=sheet_name,
            theme=theme
        )
    else:
        excel_file = create_excel_download(
            dataframe,
            filename=filename,
            sheet_name=sheet_name
        )
    
    if excel_file.getbuffer().nbytes > 0:
        st.download_button(
            label=button_label,
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.warning("Unable to create Excel file. Please check your data.")